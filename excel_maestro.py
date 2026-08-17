"""
excel_maestro.py -- Etapa 8: Exportacion del archivo final
----------------------------------------------------------------
Escribe todas las tablas Dim/Fact ya consolidadas y validadas en un unico
archivo .xlsx, con formato consistente (encabezados, anchos de columna,
tablas nombradas de Excel) para que sea directamente utilizable en Power BI.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = "Arial"
COLOR_ENCABEZADO = "0B2A4A"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _escribir_hoja(wb: Workbook, nombre_hoja: str, df: pd.DataFrame):
    ws = wb.create_sheet(nombre_hoja[:31])
    if df is None or df.empty:
        ws.append(["Sin datos disponibles todavia para esta hoja."])
        return

    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        celda = ws.cell(row=1, column=c)
        celda.font, celda.fill = HEADER_FONT, HEADER_FILL
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for _, fila in df.iterrows():
        valores = []
        for v in fila:
            if pd.isna(v) if not isinstance(v, (list, dict)) else False:
                valores.append(None)
            elif isinstance(v, bool):
                valores.append("VERDADERO" if v else "FALSO")
            else:
                valores.append(v)
        ws.append(valores)

    n_filas, n_columnas = df.shape
    for r in range(2, n_filas + 2):
        for c in range(1, n_columnas + 1):
            ws.cell(row=r, column=c).font = BODY_FONT
            ws.cell(row=r, column=c).border = BORDER

    for c in range(1, n_columnas + 1):
        letra = get_column_letter(c)
        largo = max([len(str(df.columns[c - 1]))] +
                    [len(str(x)) for x in df.iloc[:, c - 1].astype(str).tolist()[:300]])
        ws.column_dimensions[letra].width = min(max(12, largo + 2), 45)

    ws.freeze_panes = "A2"
    referencia = f"A1:{get_column_letter(n_columnas)}{n_filas + 1}"
    tabla_excel = Table(displayName=f"Tbl_{nombre_hoja.replace(' ', '_')[:25]}", ref=referencia)
    tabla_excel.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabla_excel)


def _normalizar_columnas_id(df: pd.DataFrame) -> pd.DataFrame:
    """Al releer un Excel ya guardado, pandas puede inferir columnas de
    codigo (ID_Asesor, ID_Promotor, ID_Periodo) como numericas si todos los
    valores parecen numeros, lo cual rompe comparaciones y fusiones futuras.
    Se fuerza texto de forma consistente, y se restauran a booleano real
    las columnas que se guardaron como texto VERDADERO/FALSO."""
    for col in ("ID_Asesor", "ID_Promotor", "ID_Periodo"):
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: str(int(v)) if isinstance(v, float) and not pd.isna(v) and v == int(v)
                else ("" if pd.isna(v) else str(v))
            )
    for col in df.columns:
        valores = df[col].dropna().unique()
        if len(valores) > 0 and set(map(str, valores)) <= {"VERDADERO", "FALSO"}:
            df[col] = df[col].map({"VERDADERO": True, "FALSO": False})
    return df


def cargar_excel_maestro_existente(archivo_subido, nombres_hojas: list) -> dict:
    """archivo_subido puede ser una ruta de archivo o un objeto tipo
    file-like (como el que entrega Streamlit). Regresa un diccionario
    {nombre_hoja: DataFrame}, vacio si no se proporciono archivo."""
    if archivo_subido is None:
        return {}
    xls = pd.ExcelFile(archivo_subido)
    tablas = {}
    for hoja in nombres_hojas:
        if hoja in xls.sheet_names:
            tablas[hoja] = _normalizar_columnas_id(pd.read_excel(xls, sheet_name=hoja))
    return tablas
def generar_excel_maestro(tablas: dict, ruta_salida: str):
    """tablas: diccionario {nombre_hoja: DataFrame}. El orden de insercion
    del diccionario determina el orden de las hojas en el Excel -- se
    recomienda pasar primero Resumen_Bonos e Historico_Actividad, que son
    las hojas centrales para Power BI."""
    wb = Workbook()
    wb.remove(wb.active)
    for nombre_hoja, df in tablas.items():
        _escribir_hoja(wb, nombre_hoja, df)
    Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
